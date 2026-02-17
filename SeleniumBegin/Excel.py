from openpyxl.reader.excel import load_workbook

work_book=load_workbook("C://Users//priyanshi//OneDrive//Documents//SeleniumDemo.xlsx")
sheet= work_book.active
dict= {}
print(sheet.cell(row=1, column=2))
sheet.cell(row=4, column=2).value='Aman'
sheet.cell(row=4, column=3).value='Maurya'
sheet.cell(row=4, column=4).value='Aman@gmail.com'

# print(sheet.cell(row=2, column=2).value)
print(f"{sheet.max_row} and {sheet.max_column}")
for i in range(1,sheet.max_column+1):
    print(sheet.cell(row=2, column=i).value)
 # print whole data
# for j in range(1,sheet.max_row+1):
#    for k in range(1,sheet.max_column+1):
#        print(sheet.cell(row=j, column=k).value,end='\t')
#    print()

# print specific row value based on condition
for j in range(1,sheet.max_row+1):
    if sheet.cell(row=j, column=1).value == 'Testcase2':
        for k in range(1,sheet.max_column+1):
            dict[sheet.cell(row=1,column=k).value] = sheet.cell(row=j, column=k).value
        print(dict,end='\t')
    print()



